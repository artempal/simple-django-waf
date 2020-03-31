import os
import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'waf.settings')
django.setup()
from main.models import BlackList, AttackType


def clear_all():
    BlackList.objects.all().delete()
    AttackType.objects.all().delete()


def create_attack_type():
    AttackType.objects.bulk_create([AttackType(1, "Injection"),
                                    AttackType(2, "UWA"),
                                    AttackType(3, "Scanner"),
                                    AttackType(4, "Other"),
                                    AttackType(5, "XSS"),
                                    AttackType(6, "Evasion")])


def import_table():
    workbook = load_workbook('data.xlsx', read_only=True)
    first_sheet = workbook.sheetnames[0]
    worksheet = workbook[first_sheet]

    i = 0

    attack_type = AttackType.objects
    for row in worksheet.iter_rows():
        i += 1
        if i == 1:  # пропускаем заголовок таблицы
            continue
        blacklist = BlackList()
        math_zone = row[5].value
        reg_type = row[1].value
        reg = row[2].value
        tag = row[3].value
        if math_zone.find('BODY') != -1:
            blacklist.body = True
        if math_zone.find('HEADERS') != -1:
            if math_zone.find('$HEADERS_VAR') != -1:
                blacklist.head = False
            else:
                blacklist.head = True
        if math_zone.find('URL') != -1:
            if math_zone.find('$URL') != -1:
                blacklist.url = False
            else:
                blacklist.url = True
        if math_zone.find('ARGS') != -1:
            blacklist.args = True

        blacklist.active = True

        if reg_type == 'RLx':
            blacklist.stable = False
        elif reg_type == 'RL':
            blacklist.stable = True
        blacklist.reg = reg
        blacklist.type = attack_type.get(name=tag)
        blacklist.save()


if __name__ == "__main__":
    clear_all()
    create_attack_type()
    import_table()
